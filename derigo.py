import os
import re
import sys
import json
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from datetime import datetime

from models.store import Store
from models.product import Product
from models.variant import Variant
from models.metafields import Metafields
import glob
import requests

import threading
from openpyxl import Workbook
from openpyxl.drawing.image import Image as Imag
from PIL import Image
from lxml import html

from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

import warnings
warnings.filterwarnings("ignore")

import re
import json
import threading
import requests
from lxml import html
from datetime import datetime

from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options


from models.store import Store
# from models.brand import Brand
from models.product import Product
from models.metafields import Metafields
from models.variant import Variant

from selenium.webdriver.chrome.service import Service as ChromeService
# from webdriver_manager.chrome import ChromeDriverManager

class myScrapingThread(threading.Thread):
    def __init__(self, threadID: int, name: str, obj, brand_name: str, brand_url: str, glasses_type: str, product_url: str, product_number: str) -> None:
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.brand_name = brand_name
        self.brand_url = brand_url
        self.glasses_type = glasses_type
        self.product_url = product_url
        self.product_number = product_number
        self.obj = obj
        self.status = 'in progress'
        pass

    def run(self):
        self.obj.get_product_details(self.brand_name, self.brand_url, self.glasses_type, self.product_url, self.product_number,)
        self.status = 'completed'

    def active_threads(self):
        return threading.activeCount()

class DeRigo_Scraper:
    def __init__(self, DEBUG: bool, result_filename: str, logs_filename: str, chrome_path: str) -> None:
        self.DEBUG = DEBUG
        self.data = []
        self.result_filename = result_filename
        self.logs_filename = logs_filename
        self.thread_list = []
        self.thread_counter = 0
        self.ref_json_data = None
        self.chrome_options = Options()
        self.chrome_options.add_argument('--disable-infobars')
        self.chrome_options.add_argument("--start-maximized")
        self.chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        # self.args = ["hide_console", ]
        # self.browser = webdriver.Chrome(options=self.chrome_options, service_args=self.args)
        # self.browser = webdriver.Chrome(options=self.chrome_options)
        self.browser = webdriver.Chrome(service=ChromeService(chrome_path), options=self.chrome_options)
        pass

    def controller(self, store: Store, brands_with_types: list[dict]) -> None:
        try:
            self.browser.get(store.link)
            self.wait_until_browsing()
            self.accept_cookies()

            if self.login(store.username, store.password):
                for brand_with_type in brands_with_types:
                    brand: str = brand_with_type['brand']
                    brand_code: str = str(brand_with_type['code']).strip()
                    print(f'\nBrand: {brand}')
                    self.print_logs(f'\nBrand: {brand}')

                    for glasses_type in brand_with_type['glasses_type']:
                        brand_url = self.get_brand_with_type_url(brand_code, glasses_type)
                        self.open_new_tab(brand_url)
                        
                        json_products = self.get_products_from_brand_page(brand_url)

                        total_products = len(json_products)
                        scraped_products = 0                        
                        start_time = datetime.now()
                        
                        print(f'Type: {glasses_type} | Total products: {total_products}')
                        print(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')

                        self.print_logs(f'Type: {glasses_type} | Total products: {total_products}')
                        self.print_logs(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')

                        if total_products and int(total_products) > 0: 
                            self.printProgressBar(scraped_products, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)
                        
                        for json_product in json_products:
                            scraped_products += 1
                            product_number = json_product.get('nbr')
                            product_url = json_product.get('url')
                            
                            self.get_product_details(brand, brand_url, glasses_type, product_url, product_number)
                            # self.create_thread(brand.name, brand_url, glasses_type, product_url, product_number)
                            # if self.thread_counter >= 10: 
                            #     self.wait_for_thread_list_to_complete()
                            #     self.save_to_json(self.data)

                            if total_products and int(total_products) > 0: 
                                self.printProgressBar(scraped_products, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)

                            # self.wait_for_thread_list_to_complete()
                            self.save_to_json(self.data)

                        self.close_last_tab()
                        self.save_to_json(self.data)
                        end_time = datetime.now()

                        print(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                        print('Duration: {}\n'.format(end_time - start_time))

                        self.print_logs(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                        self.print_logs('Duration: {}\n'.format(end_time - start_time))


            else: print(f'Failed to login \nURL: {store.link}\nUsername: {str(store.username)}\nPassword: {str(store.password)}')
            # input('wait')
        except Exception as e:
            self.print_logs(f'Exception in DeRigo_Scraper controller: {e}')
            if self.DEBUG: print(f'Exception in DeRigo_Scraper controller: {e}')
        finally: 
            # self.wait_for_thread_list_to_complete()
            self.save_to_json(self.data)
            self.browser.quit()

    def open_new_tab(self, url: str) -> None:
        # open category in new tab
        self.browser.execute_script('window.open("'+str(url)+'","_blank");')
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])
        self.wait_until_browsing()

    def close_last_tab(self) -> None:
        self.browser.close()
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])

    def wait_until_browsing(self) -> None:
        while True:
            try:
                state = self.browser.execute_script('return document.readyState; ')
                if 'complete' == state: break
                else: sleep(0.5)
            except: pass

    def accept_cookies(self) -> None:
        try:
            # accept cookies if found
            if self.wait_until_element_found(30, 'xpath', '//button[contains(text(), "Accept")]'):
                for _ in range(0, 20):
                    try:
                        self.browser.find_element(By.XPATH,'//button[contains(text(), "Accept")]').click()
                        sleep(0.2)
                        break
                    except: sleep(0.5)
        except Exception as e:
            self.print_logs(f'Exception in accept_cookies: {str(e)}')
            if self.DEBUG: print(f'Exception in accept_cookies: {str(e)}')
            else: pass

    def wait_until_element_found(self, wait_value: int, type: str, value: str) -> bool:
        flag = False
        try:
            if type == 'id':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.ID, value)))
                flag = True
            elif type == 'xpath':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.XPATH, value)))
                flag = True
            elif type == 'css_selector':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CSS_SELECTOR, value)))
                flag = True
            elif type == 'class_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CLASS_NAME, value)))
                flag = True
            elif type == 'tag_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.TAG_NAME, value)))
                flag = True
        except: pass
        finally: return flag

    def login(self, email: str, password: str) -> bool:
        login_flag = False
        try:
            if self.wait_until_element_found(20, 'xpath', '//input[@id="customer-account"]'):
                self.browser.find_element(By.XPATH, '//input[@id="customer-account"]').send_keys(email)
                sleep(0.2)
                if self.wait_until_element_found(20, 'xpath', '//input[@id="password"]'):
                    self.browser.find_element(By.XPATH, '//input[@id="password"]').send_keys(password)
                    sleep(0.2)
                    self.browser.find_element(By.XPATH, '//button[contains(text(), "LOG IN")]').click()

                    if self.wait_until_element_found(100, 'xpath', '//i[@class="fas fa-user"]'): login_flag = True
                else: print('Password input not found')
            else: print('Email input not found')
        except Exception as e:
            self.print_logs(f'Exception in login: {str(e)}')
            if self.DEBUG: print(f'Exception in login: {str(e)}')
            else: pass
        finally: return login_flag

    def get_brand_with_type_url(self, brand_code: str, glasses_type: str) -> str:
        brand_url: str = ''
        try:
            CodiceLivello2 = ''
            if glasses_type == 'Sunglasses': CodiceLivello2 = 'SOL'
            else: CodiceLivello2 = 'VIS'
            xpath = f'//a[contains(@href, "Linea={brand_code}") and contains(@href, "CodiceLivello2={CodiceLivello2}")]'
            
            for _ in range(0, 100):
                try:
                    brand_url = str(self.browser.find_element(By.XPATH, xpath).get_attribute('href')).strip()
                    break
                except Exception as e: 
                    if self.DEBUG: print(f'Exception in get_brand_with_type_url: {e}')
                    self.print_logs(f'Exception in get_brand_with_type_url: {e}')
                    sleep(1)
        except Exception as e:
            self.print_logs(f'Exception in get_brand_with_type_url: {str(e)}')
            if self.DEBUG: print(f'Exception in get_brand_with_type_url: {str(e)}')
            else: pass
        finally: return brand_url    

    def get_cookies(self) -> dict:
        cookies: dict = {}
        try:
            for browser_cookie in self.browser.get_cookies():
                cookies[browser_cookie['name']] = browser_cookie['value']
                # # if browser_cookie["name"] == 'php-console-server':
                # #     cookies = f'{browser_cookie["name"]}={browser_cookie["value"]}; _gat_UA-153573784-1=1; {cookies}'
                # # else:
                # cookies = f'{browser_cookie["name"]}={browser_cookie["value"]}; {cookies}'
            # cookies = cookies.strip()[:-1]
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_cookies: {e}')
            self.print_logs(f'Exception in get_cookies: {e}')
        finally: return cookies

    def get_headers(self, referer_url):
        return {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'en-US,en;q=0.9',
            'Cache-Control': 'max-age=0',
            'Connection': 'keep-alive',
            'Referer': referer_url,
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
            'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
        }

    def get_response(self, url: str, headers: dict):
        response = None
        for _ in range(0, 10):
            try:
                response = requests.get(url=url, headers=headers, cookies=self.get_cookies(), timeout=10)
                break
            except requests.exceptions.Timeout: sleep(1)
            except requests.exceptions.ConnectionError: sleep(1)
            except Exception as e:
                self.print_logs(f'Exception in get_response: {e} against {url}')
        return response
    
    def get_products_from_brand_page(self, brand_url: str) -> list[dict]:
        json_products: list[dict] = list()
        try:
            headers=self.get_headers('https://my.derigo.com/forsales2_new/index.php')
            response = self.get_response(brand_url, headers)
            if response and response.status_code == 200:
                doc_tree = html.fromstring(response.text)
                for div_tag in doc_tree.xpath('//div[@class="product" and @data-product]'):
                    json_data = json.loads(div_tag.xpath('./@data-product')[0])
                    product_number = json_data.get('id')
                    product_url = f'https://my.derigo.com/forsales2_new/articolo.php?codiceLivello1={product_number}'
                    product_json = { 'nbr': product_number, 'url': product_url }
                    if product_json not in json_products:
                        json_products.append(product_json)
                if len(doc_tree.xpath('//div[@class="product" and @data-product]')) == 12:
                    json_products += self.get_products_from_next_pages()
                
            else: self.print_logs(f'{response} for {brand_url}')
        except Exception as e:
            self.print_logs(f'Exception in get_products_from_brand_page: {str(e)}')
            if self.DEBUG: print(f'Exception in get_products_from_brand_page: {str(e)}')
            else: pass
        finally: return json_products

    def get_products_from_next_pages(self) -> list[dict]:
        json_products: list[dict] = list()
        try:
            page_products = 12
            page_no = 1
            while page_products == 12:
                page_no += 1
                next_page_url = f'https://my.derigo.com/forsales2_new/elencoProdotti.php?pag={page_no}'
                # print(next_page_url, page_products)
                headers = self.get_headers('https://my.derigo.com/forsales2_new/index.php')
                response = self.get_response(next_page_url, headers)
                
                if response and response.status_code == 200:
                    doc_tree = html.fromstring(response.text)
                    div_tags = doc_tree.xpath('//div[@class="product" and @data-product]')
                    for div_tag in div_tags:
                        json_data = json.loads(div_tag.xpath('./@data-product')[0])
                        product_number = json_data.get('id')
                        product_url = f'https://my.derigo.com/forsales2_new/articolo.php?codiceLivello1={product_number}'
                        product_json = { 'nbr': product_number, 'url': product_url }
                        if product_json not in json_products:
                            json_products.append(product_json)
                    if page_no == int(doc_tree.xpath('//select[@id="pag"]/option[@selected]/text()')[0]):
                        page_products = len(div_tags)
                    else: break
                else: 
                    self.print_logs(f'{response} for {next_page_url}')
                    break
        except Exception as e:
            self.print_logs(f'Exception in get_products_from_brand_page: {str(e)}')
            if self.DEBUG: print(f'Exception in get_products_from_brand_page: {str(e)}')
            else: pass
        finally: return json_products

    def get_product_details(self, brand_name: str, brand_url: str, glasses_type: str, product_url: str, product_number: str) -> None:
        try:
            response = requests.get(url=product_url, headers=self.get_headers(brand_url), cookies=self.get_cookies(), timeout=100)
            if response.status_code == 200:
                doc_tree = html.fromstring(response.text)

                frame_codes_and_colors = doc_tree.xpath('//div[@class="product-colors d-none d-xl-block"]/div/div/p/strong/text()')
                prices = doc_tree.xpath('//div[contains(@class, "prices-box itemModello")]/div/p[contains(text(), ",")]/text()')
                more_details = doc_tree.xpath('//div[contains(@class, "infoItems itemModello")]')
                img_urls = doc_tree.xpath('//div[contains(@class, "itemModello")]/div/a/img/@src')

                for p_index, frame_code_and_color in enumerate(frame_codes_and_colors):
                    product = Product()
                    product_price = ''

                    product.frame_code = product_number
                    product.url = product_url
                    product.brand = brand_name
                    product.type = glasses_type

                    try: product.lens_code = str(frame_code_and_color).split('-')[0].strip()
                    except: pass
                    try: product.image = str(img_urls[p_index]).strip()
                    except: pass

                    # try: product_price = str(prices[p_index]).strip()
                    # except: pass
                    
                    try: product.metafields.frame_color = str(frame_code_and_color).split('-')[-1].strip()
                    except: pass

                    for v_index, dimensions_and_availabilities in enumerate(doc_tree.xpath(f'//div[contains(@class, "color-code-list itemModello")]/ul/li[contains(@id, "rigaInput_{p_index}_")]/div[@class="infoRow row"]')):
                        variant = Variant()
                        try: variant.title = str(dimensions_and_availabilities.xpath('.//img[contains(@src, "LENTE")]/following-sibling::span/text()')[0]).strip()
                        except: pass
                        try: 
                            if not product.lens_code: variant.sku = f'{product.number} {product.frame_code} {variant.title}'
                            else: variant.sku = f'{product.number} {product.frame_code} {product.lens_code} {variant.title}'
                            if variant.sku:
                                variant.sku = str(variant.sku).strip().replace('  ', ' ')
                        except: pass
                        # if product_price: variant.listing_price = product_price
                        # if '.' in variant.listing_price: variant.listing_price = str(variant.listing_price).replace('.', '')
                        # if ',' in variant.listing_price: variant.listing_price = str(variant.listing_price).replace(',', '.')
                        # try:
                        #     match = re.search(r'([A-Z]+[0-9]+)', variant.listing_price)
                        #     if match:
                        #         variant.listing_price = match.group(1)
                        # except: pass

                        try:
                            prices = self.get_variant_price(p_index, product_url)
                            if prices:
                                variant.wholesale_price = str(prices.get('price', '')).strip()
                                variant.listing_price = str(prices.get('suggested_price', '')).strip()
                        except: pass
                        try: variant.inventory_quantity = 0 if dimensions_and_availabilities.xpath('.//div[@class="msgNoDisp"]/label[@class="rosso"]') else  5
                        except: pass

                        try: product.template = str(dimensions_and_availabilities.xpath('.//img[contains(@src, "ASTA")]/following-sibling::span/text()')[0]).strip()
                        except: pass
                        try: product.bridge = str(dimensions_and_availabilities.xpath('.//img[contains(@src, "PONTE")]/following-sibling::span/text()')[0]).strip()
                        except: pass

                        try: variant.size = f'{variant.title}-{product.bridge}-{product.template}'
                        except: pass

                        try: variant.barcode_or_gtin = str(more_details[p_index].xpath(f'.//div[contains(@id, "rigaInfo_{p_index}_{v_index}")]')[0].xpath('.//span[contains(text(), "EAN/UPC")]/parent::p/text()')[0]).replace(':', '').strip()
                        except: pass
                        product.add_single_variant(variant)

                    for variant in product.variants:
                        if variant.size: product.metafields.size_bridge_template += f'{variant.size}, '
                        if variant.barcode_or_gtin: product.metafields.gtin1 += f'{variant.barcode_or_gtin}, '
                    
                    product.metafields.for_who = 'Unisex'
                    
                    if product.metafields.size_bridge_template:
                        product.metafields.size_bridge_template = str(product.metafields.size_bridge_template).strip()
                        if product.metafields.size_bridge_template[-1] == ',': product.metafields.size_bridge_template = product.metafields.size_bridge_template[:-1]
                    if product.metafields.gtin1:
                        product.metafields.gtin1 = str(product.metafields.gtin1).strip()
                        if product.metafields.gtin1[-1] == ',': product.metafields.gtin1 = product.metafields.gtin1[:-1]


                    try: product.metafields.frame_material = str(more_details[p_index].xpath('.//span[contains(text(), "Material")]/parent::p/text()')[0]).replace(':', '').strip()
                    except: pass
                    try: product.metafields.lens_material = str(more_details[p_index].xpath('.//span[contains(text(), "Lens material")]/parent::p/text()')[0]).replace(':', '').strip()
                    except: pass
                    try: product.metafields.lens_color = str(more_details[p_index].xpath('.//span[contains(text(), "Lens colour")]/parent::p/text()')[0]).replace(':', '').strip()
                    except: pass
                    try: product.metafields.lens_technology = str(more_details[p_index].xpath('.//span[contains(text(), "Sub-group")]/parent::p/text()')[0]).replace(':', '').strip()
                    except: pass

                    self.data.append(product)
                    self.save_to_json(self.data)

        except Exception as e:
            self.print_logs(f'Exception in get_product_details: {str(e)}')
            if self.DEBUG: print(f'Exception in get_product_details: {str(e)}')
            else: pass 

    def get_variant_price(self, index: int, product_url: str) -> dict:
        prices: dict = {}
        try:
            headers=self.get_headers(product_url)

            data = {
                'flag': 'getInfoRowGtm',
                'id': str(index),
            }
            for _ in range(0, 10):
                try:
                    response = requests.post('https://my.derigo.com/forsales2_new/remoto.php', cookies=self.get_cookies(), headers=headers, data=data, timeout=100)
                    if response.status_code == 200:
                        prices = response.json()
                        break
                    else: self.print_logs(f'{response} for getting variant price from {product_url} with index {index}')
                except requests.exceptions.Timeout: sleep(1)
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_variant_price: {e}')
            self.print_logs(f'Exception in get_variant_price: {e}')
        finally: return prices

    def save_to_json(self, products: list[Product]) -> None:
        try:
            json_products = []
            for product in products:
                _id = ''
                if product.lens_code: _id = f"{str(product.number).strip().upper()}_{str(product.frame_code).strip().upper()}_{str(product.lens_code).strip().upper()}"
                else: _id = f"{str(product.number).strip().upper()}_{str(product.frame_code).strip().upper()}"
                if _id[0] == '_': _id = _id[1:]
                json_varinats = []
                for variant in product.variants:
                    json_varinat = {
                        "_id": str(variant.sku).strip().upper().replace(' ', '_'),
                        "product_id": _id,
                        'title': str(variant.title).strip(),
                        'sku': str(variant.sku).strip().upper(),
                        'inventory_quantity': int(variant.inventory_quantity),
                        'found_status': int(variant.found_status),
                        'wholesale_price': float(variant.wholesale_price) if variant.wholesale_price else 0.00,
                        'listing_price': float(variant.listing_price) if variant.listing_price else 0.00,
                        'barcode_or_gtin': str(variant.barcode_or_gtin).strip(),
                        'size': str(variant.size).strip().replace(' ', '')
                    }
                    json_varinats.append(json_varinat)


                json_product = {
                    "_id": _id,
                    'number': str(product.number).strip().upper(),
                    'name': str(product.name).strip().title(),
                    'brand': str(product.brand).strip().title(),
                    'frame_code': str(product.frame_code).strip().upper(),
                    'lens_code': product.lens_code,
                    'type': product.type,
                    'bridge': product.bridge,
                    'template': product.template,
                    "url": product.url,
                    'metafields': {
                        'for_who': str(product.metafields.for_who).strip().title(),
                        'lens_material': str(product.metafields.lens_material).strip().title(),
                        'lens_technology': str(product.metafields.lens_technology).strip().title(),
                        'lens_color': str(product.metafields.lens_color).strip().title(),
                        'frame_shape': str(product.metafields.frame_shape).strip().title(),
                        'frame_material': str(product.metafields.frame_material).strip().title(),
                        'frame_color': str(product.metafields.frame_color).strip().title(),
                        'size-bridge-template': str(product.metafields.size_bridge_template).strip(),
                        'gtin1': str(product.metafields.gtin1).strip()
                    },
                    'image': str(product.image).strip(),
                    'images_360': product.images_360,
                    'variants': json_varinats
                }
                json_products.append(json_product)


            with open(self.result_filename, 'w') as f: json.dump(json_products, f)

        except Exception as e:
            if self.DEBUG: print(f'Exception in save_to_json: {e}')
            self.print_logs(f'Exception in save_to_json: {e}')

    def create_thread(self, brand_name: str, brand_url: str, glasses_type: str, product_url: str, product_number: str) -> None:
        thread_name = "Thread-"+str(self.thread_counter)
        self.thread_list.append(myScrapingThread(self.thread_counter, thread_name, self, brand_name, brand_url, glasses_type, product_url, product_number))
        self.thread_list[self.thread_counter].start()
        self.thread_counter += 1

    def is_thread_list_complted(self) -> bool:
        for obj in self.thread_list:
            if obj.status == "in progress":
                return False
        return True

    def wait_for_thread_list_to_complete(self) -> None:
        while True:
            result = self.is_thread_list_complted()
            if result:
                self.thread_counter = 0
                self.thread_list.clear()
                break
            else: sleep(1)

    # print logs to the log file
    def print_logs(self, log: str) -> None:
        try:
            with open(self.logs_filename, 'a') as f:
                f.write(f'\n{log}')
        except: pass

    def printProgressBar(self, iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r") -> None:
        """
        Call in a loop to create terminal progress bar
        @params:
            iteration   - Required  : current iteration (Int)
            total       - Required  : total iterations (Int)
            prefix      - Optional  : prefix string (Str)
            suffix      - Optional  : suffix string (Str)
            decimals    - Optional  : positive number of decimals in percent complete (Int)
            length      - Optional  : character length of bar (Int)
            fill        - Optional  : bar fill character (Str)
            printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
        """
        percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
        filledLength = int(length * iteration // total)
        bar = fill * filledLength + '-' * (length - filledLength)
        print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
        # Print New Line on Complete
        if iteration == total:
            print()


def read_data_from_json_file(DEBUG, result_filename: str):
    data = []
    try:
        files = glob.glob(result_filename)
        if files:
            f = open(files[-1])
            json_data = json.loads(f.read())
            products = []

            for json_d in json_data:
                number, frame_code, brand, img_url, frame_color, lens_color = '', '', '', '', '', ''
                brand = json_d['brand']
                number = str(json_d['name']).strip().upper()
                if '-' in number: number = number.replace('-', '/').strip()
                frame_code = str(json_d['frame_code']).strip().upper()
                frame_color = str(json_d.get('metafields', {}).get('frame_color', '')).strip().title()
                lens_color = str(json_d.get('metafields', {}).get('lens_color', '')).strip().title()
                img_url = str(json_d.get('image', '')).strip()

                for json_variant in json_d['variants']:
                    sku, price = '', ''
                    sku = str(json_variant['sku']).strip().upper()
                    if '/' in sku: sku = sku.replace('/', '-').strip()
                    wholesale_price = str(json_variant['wholesale_price']).strip()
                    listing_price = str(json_variant['listing_price']).strip()
                    barcode_or_gtin = str(json_variant['barcode_or_gtin']).strip()
                    image_filname = f'Images/{sku}.jpg'
                    # if not os.path.exists(image_filname):
                    #     image_attachment = download_image(img_url)
                    #     if image_attachment:
                    #         with open(f'Images/{sku}.jpg', 'wb') as f: f.write(image_attachment)
                    #         crop_downloaded_image(f'Images/{sku}.jpg')

                    data.append([brand, number, frame_code, frame_color, lens_color,  sku, wholesale_price, listing_price, barcode_or_gtin])
    except Exception as e:
        if DEBUG: print(f'Exception in read_data_from_json_file: {e}')
        else: pass
    finally: return data

def download_image(url):
    image_attachment = ''
    try:
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-Encoding': 'gzip, deflate, br',
            'accept-Language': 'en-US,en;q=0.9',
            'cache-Control': 'max-age=0',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'Sec-Fetch-User': '?1',
            'upgrade-insecure-requests': '1',
        }
        counter = 0
        while True:
            try:
                response = requests.get(url=url, headers=headers, timeout=20)
                # print(response.status_code)
                if response.status_code == 200:
                    # image_attachment = base64.b64encode(response.content)
                    image_attachment = response.content
                    break
                else: print(f'{response.status_code} found for downloading image')
            except: sleep(0.3)
            counter += 1
            if counter == 10: break
    except Exception as e: print(f'Exception in download_image: {str(e)}')
    finally: return image_attachment

def crop_downloaded_image(filename):
    try:
        im = Image.open(filename)
        width, height = im.size   # Get dimensions
        new_width = 1120
        new_height = 600
        if width > new_width and height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            im.save(filename)
        elif height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            im.save(filename)
    except Exception as e: print(f'Exception in crop_downloaded_image: {e}')

def saving_picture_in_excel(data: list):
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value='Brand')
    worksheet.cell(row=1, column=2, value='Model Code')
    worksheet.cell(row=1, column=3, value='Lens Code')
    worksheet.cell(row=1, column=4, value='Color Frame')
    worksheet.cell(row=1, column=5, value='Color Lens')
    worksheet.cell(row=1, column=6, value='SKU')
    worksheet.cell(row=1, column=7, value='Wholesale Price')
    worksheet.cell(row=1, column=8, value='Listing Price')
    worksheet.cell(row=1, column=9, value="UPC")
    worksheet.cell(row=1, column=10, value="Image")

    for index, d in enumerate(data):
        new_index = index + 2

        worksheet.cell(row=new_index, column=1, value=d[0])
        worksheet.cell(row=new_index, column=2, value=d[1])
        worksheet.cell(row=new_index, column=3, value=d[2])
        worksheet.cell(row=new_index, column=4, value=d[3])
        worksheet.cell(row=new_index, column=5, value=d[4])
        worksheet.cell(row=new_index, column=6, value=d[5])
        worksheet.cell(row=new_index, column=7, value=d[6])
        worksheet.cell(row=new_index, column=8, value=d[7])
        worksheet.cell(row=new_index, column=9, value=d[8])

        image = f'Images/{d[-4]}.jpg'
        if os.path.exists(image):
            im = Image.open(image)
            width, height = im.size
            worksheet.row_dimensions[new_index].height = height
            worksheet.add_image(Imag(image), anchor='J'+str(new_index))
            # col_letter = get_column_letter(7)
            # worksheet.column_dimensions[col_letter].width = width

    workbook.save('Derigo Results.xlsx')

DEBUG = True
try:
    pathofpyfolder = os.path.realpath(sys.argv[0])
    # get path of Exe folder
    path = pathofpyfolder.replace(pathofpyfolder.split('\\')[-1], '')
    
    if os.path.exists('Derigo Results.xlsx'): os.remove('Derigo Results.xlsx')

    if '.exe' in pathofpyfolder.split('\\')[-1]: DEBUG = False
    
    f = open('Derigo start.json')
    json_data = json.loads(f.read())
    f.close()

    brands = json_data['brands']

    
    f = open('requirements/derigo.json')
    data = json.loads(f.read())
    f.close()

    store = Store()
    store.link = data['url']
    store.username = data['username']
    store.password = data['password']
    store.login_flag = True

    result_filename = 'requirements/Derigo Results.json'

    if not os.path.exists('Logs'): os.makedirs('Logs')

    log_files = glob.glob('Logs/*.txt')
    if len(log_files) > 5:
        oldest_file = min(log_files, key=os.path.getctime)
        os.remove(oldest_file)
        log_files = glob.glob('Logs/*.txt')

    scrape_time = datetime.now().strftime('%d-%m-%Y %H-%M-%S')
    logs_filename = f'Logs/Logs {scrape_time}.txt'

    chrome_path = ''
    if not chrome_path:
        chrome_path = ChromeDriverManager().install()
        if 'chromedriver.exe' not in chrome_path:
            chrome_path = str(chrome_path).split('/')[0].strip()
            chrome_path = f'{chrome_path}\\chromedriver.exe'
    
    # DeRigo_Scraper(DEBUG, result_filename, logs_filename, chrome_path).controller(store, brands)
    
    # for filename in glob.glob('Images/*'): os.remove(filename)
    data = read_data_from_json_file(DEBUG, result_filename)
    # os.remove(result_filename)

    saving_picture_in_excel(data)
except Exception as e:
    if DEBUG: print('Exception: '+str(e))
    else: pass
